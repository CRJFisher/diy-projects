#!/usr/bin/env python3
"""
Final pre-B&Q shopping list update — 2026-04-11

Changes made and why:
1. 50x47mm 2.4m qty: 10 -> 8
   Braces moved to user-supplied 15x38mm, freeing 2 sticks.
   Cutting plan confirms 8 sticks needed (6 posts + 2 door stiles, each paired with 750mm offcut).

2. 50x47mm 1.8m qty: 8 -> 5
   Noggins removed; only 5 sticks needed (4x 1300mm rails + 1x 500mm door rail + 1x 423mm front bottom rail).
   Stick 1 pairs a 1300mm rail with the 423mm rail. Sticks 2-4 each hold one 1300mm rail. Stick 5 holds the 500mm door rail.

3. Screw notes updated:
   Old: ~56x 75mm, ~18x 65mm, ~24x 50mm (noggins) = ~98 total
   New: ~40x 75mm (joints), ~16x 65mm (braces) = ~56 total. No 50mm screws needed (noggins removed).

4. Added note: X-braces are user-supplied 15x38mm timber, NOT from B&Q.

5. Transport note updated: offcuts from 2.4m cuts are 750mm depth rails (no longer braces).

6. Verified: 3x extra 50x22mm 1.8m sticks (item 5) and 1x plywood sheet (item 6) are present.
"""

import openpyxl

path = 'Weekend 1 Shopping List.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# --- Change 1: 2.4m sticks qty 10 -> 8 ---
assert ws['D7'].value == 10, f"Expected D7=10, got {ws['D7'].value}"
ws['D7'] = 8
print("D7 (2.4m qty): 10 -> 8")

# --- Change 2: 1.8m sticks qty 8 -> 5 ---
assert ws['D8'].value == 8, f"Expected D8=8, got {ws['D8'].value}"
ws['D8'] = 5
print("D8 (1.8m qty): 8 -> 5")

# --- Change 3: Update screw notes ---
old_screw_note = ws['H12'].value
new_screw_note = (
    "Pick from the screws aisle \u2014 look for outdoor/galvanised.\n"
    "Weekend 1 needs: ~40x 75mm (joints), ~16x 65mm (braces) = ~56 total.\n"
    "No 50mm screws needed (noggins removed from design).\n"
    "A 300-500 pc assorted box covers all weekends (~200-300 total)."
)
ws['H12'] = new_screw_note
print("H12 (screw notes): updated counts (56 total, no 50mm)")

# --- Change 4: Add X-brace user-supplied note ---
# Insert a note row after the fixings section (after row 12) but before additional items.
# Actually, better to add it as a note in a new row. Let's put it after the fixings row.
# Current layout: row 12 = screws, row 13 = "ADDITIONAL ITEMS" header.
# Insert a new row at 13 to hold the brace note.
ws.insert_rows(13)
ws['A13'] = 'NOTE'
ws['B13'] = (
    'X-braces: user-supplied 15x38x2400mm timber (NOT purchased from B&Q). '
    'Only 2 of 8 available sticks needed for 4 braces. 6 sticks spare.'
)
# Style it like a section header
from openpyxl.styles import Font
ws['A13'].font = Font(bold=True, color='666666')
ws['B13'].font = Font(italic=True, color='666666')
print("Row 13: inserted X-brace user-supplied note")

# --- Change 5: Update transport note about offcuts ---
# The transport note was in row 20 (now row 21 after insert).
transport_row = 21
old_transport = ws.cell(row=transport_row, column=3).value
new_transport = (
    "The 2.4m posts won't fit in the Honda Insight (max ~1.8m with seats down). "
    "SOLUTION: Ask B&Q to make ONE cut per stick at the post length (e.g. 1647mm). "
    "KEEP BOTH PIECES \u2014 the offcuts are exactly 750mm (depth rails). "
    "Without these offcuts you would need ~7 extra 1.8m sticks. "
    "Options if they won't cut: (1) Arrange delivery. (2) Hire a van for ~\u00a330."
)
ws.cell(row=transport_row, column=3).value = new_transport
print(f"Row {transport_row}: updated transport note (offcuts = depth rails only)")

# --- Change 6: Update verified date ---
ws['A3'] = 'VERIFIED against diy.com and cutting plan on 11 April 2026 \u2014 FINAL pre-B&Q check'
ws['A3'].font = Font(bold=True)
print("A3: updated verification date to final check")

wb.save(path)
print(f"\nSaved: {path}")
print("\n--- Summary of quantities ---")
print(f"  50x47mm 2.4m sticks: 8  (was 10, saved 2)")
print(f"  50x47mm 1.8m sticks: 5  (was 8, saved 3)")
print(f"  50x22mm 1.8m battens (roof): 4  (unchanged)")
print(f"  50x22mm 1.8m battens (shelf, buy-ahead): 3  (unchanged)")
print(f"  Plywood sheet: 1  (unchanged)")
print(f"  Screw box: 1  (unchanged, but notes updated)")
print(f"  15x38mm braces: NOT on list (user-supplied, noted)")
