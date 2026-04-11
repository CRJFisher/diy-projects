#!/usr/bin/env python3
"""
Update Weekend 1 Shopping List.xlsx
====================================
Ensures the spreadsheet reflects the verified Weekend 1 materials:
- Correct screw sizes in fixings description (50mm, 65mm, 75mm)
- Accurate screw count guidance in notes
- All quantities verified against cutting plan and weekend-1-guide.md

Run: python3 update_weekend1_sheet.py
"""

import openpyxl

SPREADSHEET = "/Users/chuck/workspace/diy-projects/Weekend 1 Shopping List.xlsx"

wb = openpyxl.load_workbook(SPREADSHEET)
ws = wb.active

changes = []

# ── 1. Fix screw description (B12): "40mm + 75mm" → "50mm, 65mm, 75mm" ──
old_b12 = ws["B12"].value
new_b12 = old_b12.replace("40mm + 75mm for frame", "50mm, 65mm, 75mm for frame")
if old_b12 != new_b12:
    ws["B12"].value = new_b12
    changes.append(f"B12: Updated screw sizes from '40mm + 75mm' to '50mm, 65mm, 75mm'")
else:
    changes.append("B12: Already correct (no change)")

# ── 2. Update screw notes (H12) with verified counts ──
old_h12 = ws["H12"].value
new_h12 = (
    "Pick from the screws aisle — look for outdoor/galvanised.\n"
    "Weekend 1 needs: ~40x 75mm (joints), ~16x 65mm (braces) = ~56 total.\n"
    "A 300-500 pc assorted box covers all weekends (~200-300 total)."
)
if old_h12 != new_h12:
    ws["H12"].value = new_h12
    changes.append(f"H12: Updated screw notes with verified counts")
else:
    changes.append("H12: Already correct (no change)")

# ── 3. Verify timber quantities (should be 10x 2.4m, 8x 1.8m) ──
d7 = ws["D7"].value  # 2.4m sticks
d8 = ws["D8"].value  # 1.8m sticks
d10 = ws["D10"].value  # 50x22mm battens
d14 = ws["D14"].value  # extra shelf battens
d15 = ws["D15"].value  # plywood

assert d7 == 10, f"Expected 10x 2.4m sticks, got {d7}"
assert d8 == 8, f"Expected 8x 1.8m sticks, got {d8}"
assert d10 == 4, f"Expected 4x 50x22mm roof battens, got {d10}"
assert d14 == 3, f"Expected 3x 50x22mm shelf battens, got {d14}"
assert d15 == 1, f"Expected 1x plywood sheet, got {d15}"
changes.append("D7/D8/D10/D14/D15: All quantities verified correct")

# ── 4. Check transport note doesn't reference stale 470mm ──
c20 = ws["C20"].value or ""
if "470mm" in c20:
    ws["C20"].value = c20.replace("470mm", "423mm (right section clear span)")
    changes.append("C20: Fixed stale 470mm reference to 423mm")
else:
    changes.append("C20: No stale 470mm reference (ok)")

# ── Save ──
wb.save(SPREADSHEET)

print("=" * 60)
print("Weekend 1 Shopping List — Update Report")
print("=" * 60)
for c in changes:
    print(f"  {c}")
print()
print(f"Saved to: {SPREADSHEET}")
